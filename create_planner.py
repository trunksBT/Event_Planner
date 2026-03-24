import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============================================================
# CONFIGURATION — edit these values to customize the planner
# ============================================================
YEAR = 2026

MONTHS = [
    (8, "August", [31]),     # 3rd element: days from previous month to include
    (9, "September", []),
]

PEOPLE = [
    "Alice", "Bob", "Charlie", "Diana", "Eve",
    "Frank", "Grace", "Hank", "Ivy", "Jack",
    "Karen", "Leo", "Mia", "Noah", "Olivia",
    "Paul", "Quinn", "Rita", "Sam", "Tina",
]

CRITICAL = ["Alice", "Bob", "Diana", "Frank"]

# Weekend period — which weekdays count as "weekend" for the event?
#   0=Mon, 1=Tue, 2=Wed, 3=Thu, 4=Fri, 5=Sat, 6=Sun
# Examples:
#   Sat-Sun only     → [5, 6]
#   Fri-Sun          → [4, 5, 6]
#   Fri-Mon (long)   → [4, 5, 6, 0]
WEEKEND_DAYS = [4, 5, 6]   # ← Friday to Sunday

UNAVAILABLE_MARK = "X"

# ============================================================
# STYLES
# ============================================================
DAY_ABBR = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

def weekend_label():
    names = [DAY_ABBR[d] for d in sorted(WEEKEND_DAYS)]
    return f"{names[0]}-{names[-1]}"

white_sm = Font(name="Arial", bold=True, size=9, color="FFFFFF")
white_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
normal_font = Font(name="Arial", size=10)
critical_font = Font(name="Arial", bold=True, size=10, color="B8860B")
title_font = Font(name="Arial", bold=True, size=14, color="1F3864")
subtitle_font = Font(name="Arial", italic=True, size=10, color="555555")
red_x_font = Font(name="Arial", bold=True, size=12, color="C00000")
green_font = Font(name="Arial", bold=True, size=11, color="1B7A2B")

fill_hdr = PatternFill("solid", fgColor="2F5496")
fill_wknd_hdr = PatternFill("solid", fgColor="1B3A6B")
fill_wknd_col = PatternFill("solid", fgColor="E8EDF5")
fill_critical_row = PatternFill("solid", fgColor="FFF8E1")
fill_crit_wknd = PatternFill("solid", fgColor="FFF3D6")
fill_legend_bg = PatternFill("solid", fgColor="F2F2F2")
fill_unavail = PatternFill("solid", fgColor="FFCDD2")
fill_crit_marker = PatternFill("solid", fgColor="FFD966")
fill_result_hdr = PatternFill("solid", fgColor="4472C4")
fill_best = PatternFill("solid", fgColor="C6EFCE")
fill_second_best = PatternFill("solid", fgColor="70AD47")
fill_wknd_bad = PatternFill("solid", fgColor="FFC7CE")

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_a = Alignment(horizontal="left", vertical="center")
thin = Border(
    left=Side("thin", color="B0B0B0"), right=Side("thin", color="B0B0B0"),
    top=Side("thin", color="B0B0B0"), bottom=Side("thin", color="B0B0B0"),
)

# ============================================================
# BUILD EACH SHEET
# ============================================================
for idx, (month_num, month_name, prev_days) in enumerate(MONTHS):
    ws = wb.active if idx == 0 else wb.create_sheet()
    ws.title = month_name
    num_days = calendar.monthrange(YEAR, month_num)[1]

    # Build full list of days: optional previous-month days + current month
    prev_month = month_num - 1 if month_num > 1 else 12
    sheet_days = [(prev_month, d) for d in prev_days] + \
                 [(month_num, d) for d in range(1, num_days + 1)]
    total_cols = len(sheet_days)

    DATA_COL_START = 3   # col C = first day

    ws.freeze_panes = "C5"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 11
    for di in range(total_cols):
        ws.column_dimensions[get_column_letter(DATA_COL_START + di)].width = 6.5

    # ---------- ROW 1 — Title ----------
    last_col = DATA_COL_START + total_cols - 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    tc = ws.cell(row=1, column=1,
                 value=f"Event Availability Planner  —  {month_name} {YEAR}")
    tc.font = title_font
    tc.alignment = left_a
    ws.row_dimensions[1].height = 32

    # ---------- ROW 2 — Subtitle ----------
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    sc = ws.cell(row=2, column=1,
                 value=f'Weekend: {weekend_label()}  |  '
                       f'Mark "{UNAVAILABLE_MARK}" = unavailable  |  '
                       f'\u2605 = critical member')
    sc.font = subtitle_font
    ws.row_dimensions[2].height = 20

    # ---------- ROW 3 — Weekday name headers ----------
    for c, (val, fll) in enumerate([
        ("Person", fill_hdr), ("Critical?", fill_crit_marker)
    ], start=1):
        cell = ws.cell(row=3, column=c, value=val)
        cell.font = white_sm if c == 1 else Font(name="Arial", bold=True, size=9)
        cell.fill = fll
        cell.alignment = center
        cell.border = thin

    for di, (dm, dd) in enumerate(sheet_days):
        col = DATA_COL_START + di
        dow = calendar.weekday(YEAR, dm, dd)
        cell = ws.cell(row=3, column=col, value=DAY_ABBR[dow])
        cell.font = white_sm
        cell.fill = fill_wknd_hdr if dow in WEEKEND_DAYS else fill_hdr
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[3].height = 18

    # ---------- ROW 4 — Day-number headers ----------
    ws.cell(row=4, column=1).fill = fill_hdr
    ws.cell(row=4, column=1).border = thin
    ws.cell(row=4, column=2).fill = fill_crit_marker
    ws.cell(row=4, column=2).border = thin

    for di, (dm, dd) in enumerate(sheet_days):
        col = DATA_COL_START + di
        dow = calendar.weekday(YEAR, dm, dd)
        cell = ws.cell(row=4, column=col, value=dd)
        cell.font = white_font
        cell.fill = fill_wknd_hdr if dow in WEEKEND_DAYS else fill_hdr
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[4].height = 22

    # ---------- PEOPLE ROWS (row 5+) ----------
    FIRST_P = 5
    for pi, person in enumerate(PEOPLE):
        row = FIRST_P + pi
        is_crit = person in CRITICAL

        nc = ws.cell(row=row, column=1, value=person)
        nc.font = normal_font
        nc.alignment = left_a
        nc.border = thin

        cc = ws.cell(row=row, column=2, value="\u2605" if is_crit else "")
        cc.font = critical_font
        cc.alignment = center
        cc.border = thin

        for di, (dm, dd) in enumerate(sheet_days):
            col = DATA_COL_START + di
            dow = calendar.weekday(YEAR, dm, dd)
            cell = ws.cell(row=row, column=col)
            cell.alignment = center
            cell.font = normal_font
            cell.border = thin
            if dow in WEEKEND_DAYS:
                cell.fill = fill_wknd_col

        ws.row_dimensions[row].height = 22

    LAST_P = FIRST_P + len(PEOPLE) - 1

    # ---------- ANALYSIS SECTION ----------
    A_TITLE = LAST_P + 2
    R_UNAVAIL = A_TITLE + 1
    R_CRIT_UN = R_UNAVAIL + 1
    R_FREE    = R_CRIT_UN + 1
    R_BEST    = R_FREE + 1
    R_CRIT_WK = R_BEST + 1   # hidden: critical unavail sum per weekend column
    R_HELPER  = R_BEST + 2   # hidden: total unavail sum per group (for ranking)

    # Analysis header bar
    ws.merge_cells(start_row=A_TITLE, start_column=1, end_row=A_TITLE, end_column=2)
    ah = ws.cell(row=A_TITLE, column=1, value="ANALYSIS")
    ah.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ah.fill = fill_result_hdr
    ah.alignment = center
    ah.border = thin
    ws.cell(row=A_TITLE, column=2).fill = fill_result_hdr
    ws.cell(row=A_TITLE, column=2).border = thin

    # Row labels
    for r, lbl in [
        (R_UNAVAIL, "Unavailable total"),
        (R_CRIT_UN, "Critical unavail"),
        (R_FREE,    "All critical free?"),
        (R_BEST,    "Best weekend?"),
    ]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws.cell(row=r, column=1, value=lbl)
        c.font = Font(name="Arial", bold=True, size=10)
        c.alignment = left_a
        c.fill = fill_legend_bg
        c.border = thin
        ws.cell(row=r, column=2).fill = fill_legend_bg
        ws.cell(row=r, column=2).border = thin

    # Range references for dynamic critical formulas (reads ★ from column B)
    crit_range = f"$B${FIRST_P}:$B${LAST_P}"

    # Group weekend days into whole weekends (e.g., Fri+Sat+Sun = one weekend)
    wknd_groups = []      # list of lists of column letters
    di_to_group = {}      # sheet_days index → group index
    current_group = []
    for di, (dm, dd) in enumerate(sheet_days):
        if calendar.weekday(YEAR, dm, dd) in WEEKEND_DAYS:
            current_group.append(get_column_letter(DATA_COL_START + di))
            di_to_group[di] = len(wknd_groups)
        else:
            if current_group:
                wknd_groups.append(current_group)
                current_group = []
    if current_group:
        wknd_groups.append(current_group)
    # Sum of unavailable / critical unavailable per weekend group
    group_sums = []
    crit_group_sums = []
    for group in wknd_groups:
        group_sums.append("+".join(f"{cl}{R_UNAVAIL}" for cl in group))
        crit_group_sums.append("+".join(f"{cl}{R_CRIT_UN}" for cl in group))
    for di, (dm, dd) in enumerate(sheet_days):
        col = DATA_COL_START + di
        cl = get_column_letter(col)
        dow = calendar.weekday(YEAR, dm, dd)
        is_wknd = dow in WEEKEND_DAYS

        # Extend the header bar across
        ws.cell(row=A_TITLE, column=col).fill = fill_result_hdr
        ws.cell(row=A_TITLE, column=col).border = thin

        # Total unavailable
        f1 = f'=COUNTIF({cl}{FIRST_P}:{cl}{LAST_P},"{UNAVAILABLE_MARK}")'
        c1 = ws.cell(row=R_UNAVAIL, column=col, value=f1)
        c1.font = normal_font; c1.alignment = center; c1.border = thin
        if is_wknd: c1.fill = fill_wknd_col

        # Critical unavailable (dynamic — counts people with ★ in col B and X in this day)
        day_range = f"{cl}{FIRST_P}:{cl}{LAST_P}"
        f2 = f'=SUMPRODUCT(({crit_range}="\u2605")*({day_range}="{UNAVAILABLE_MARK}"))'
        c2 = ws.cell(row=R_CRIT_UN, column=col, value=f2)
        c2.font = normal_font; c2.alignment = center; c2.border = thin
        if is_wknd: c2.fill = fill_wknd_col

        # All critical free?
        crit_cell = f"{cl}{R_CRIT_UN}"
        f3 = f'=IF({crit_cell}=0,"\u2713","\u2717")'
        c3 = ws.cell(row=R_FREE, column=col, value=f3)
        c3.font = green_font; c3.alignment = center; c3.border = thin
        if is_wknd: c3.fill = fill_wknd_col

        # Best weekend — cell holds the weekend group sum (hidden by ;;; format);
        # conditional formatting colors best/2nd best (green) or critical-blocked (red)
        if is_wknd:
            gi = di_to_group[di]
            c4 = ws.cell(row=R_BEST, column=col, value=f"={group_sums[gi]}")
            c4.number_format = ';;;'
            c4.alignment = center; c4.border = thin; c4.fill = fill_wknd_col
            # Hidden row: critical unavailable sum for this weekend
            ck = ws.cell(row=R_CRIT_WK, column=col, value=f"={crit_group_sums[gi]}")
            ck.number_format = ';;;'
        else:
            c4 = ws.cell(row=R_BEST, column=col)
            c4.fill = fill_legend_bg; c4.border = thin

    for gi in range(len(wknd_groups)):
        hcol = DATA_COL_START + gi
        ws.cell(row=R_HELPER, column=hcol, value=f"={group_sums[gi]}").number_format = ';;;'
    ws.row_dimensions[R_CRIT_WK].hidden = True
    ws.row_dimensions[R_HELPER].hidden = True

    # Conditional formatting on R_BEST row (color = rank)
    h_start = f"${get_column_letter(DATA_COL_START)}${R_HELPER}"
    h_end = f"${get_column_letter(DATA_COL_START + len(wknd_groups) - 1)}${R_HELPER}"
    h_range = f"{h_start}:{h_end}"
    cf_first = get_column_letter(DATA_COL_START)
    cf_last = get_column_letter(last_col)
    cf_range = f"{cf_first}{R_BEST}:{cf_last}{R_BEST}"
    ref = f"{cf_first}{R_BEST}"
    # Second-best = SMALL(helper, k) where k skips all cells tied for MIN
    second_val = f'SMALL({h_range},COUNTIF({h_range},MIN({h_range}))+1)'
    crit_ref = f"{cf_first}{R_CRIT_WK}"  # same column, shifts with CF range
    # Rule 1 (highest priority): critical person unavailable → red
    ws.conditional_formatting.add(cf_range, FormulaRule(
        formula=[f'AND(ISNUMBER({ref}),{crit_ref}>0)'],
        fill=fill_wknd_bad, stopIfTrue=True,
    ))
    # Rule 2: best weekend → light green
    ws.conditional_formatting.add(cf_range, FormulaRule(
        formula=[f'AND(ISNUMBER({ref}),{ref}=MIN({h_range}))'],
        fill=fill_best, stopIfTrue=True,
    ))
    # Rule 3: second best → dark green
    ws.conditional_formatting.add(cf_range, FormulaRule(
        formula=[f'AND(ISNUMBER({ref}),IFERROR({second_val},FALSE)>MIN({h_range}),{ref}=IFERROR({second_val},FALSE))'],
        fill=fill_second_best,
    ))

    # ---------- LEGEND ----------
    LR = R_BEST + 5
    ws.merge_cells(start_row=LR, start_column=1, end_row=LR, end_column=7)
    lg = ws.cell(row=LR, column=1, value="LEGEND")
    lg.font = Font(name="Arial", bold=True, size=11)
    lg.fill = fill_legend_bg; lg.border = thin

    legend = [
        (LR+1, f'"{UNAVAILABLE_MARK}"', "Person is NOT available on this day",
         fill_unavail, red_x_font),
        (LR+2, "(empty)", "Person IS available (or hasn't responded yet)",
         None, normal_font),
        (LR+3, "\u2605",
         'Critical person — add/remove \u2605 in the "Critical?" column to update analysis',
         fill_crit_marker, critical_font),
        (LR+4, "Shaded cols", f"Weekend columns ({weekend_label()}) — highlighted for scanning",
         fill_wknd_col, normal_font),
    ]
    # Analysis explanation entries
    legend += [
        (LR+6, "ANALYSIS", "", fill_result_hdr,
         Font(name="Arial", bold=True, size=10, color="FFFFFF")),
        (LR+7, "Unavail total",
         f'Counts how many people marked "{UNAVAILABLE_MARK}" on each day',
         fill_legend_bg, normal_font),
        (LR+8, "Critical unavail",
         'Counts unavailable people who have \u2605 in column B (editable in Excel)',
         fill_legend_bg, normal_font),
        (LR+9, "\u2713 / \u2717",
         "All critical free? \u2713 = no critical person is unavailable, \u2717 = at least one is",
         fill_legend_bg, green_font),
        (LR+10, "Light green",
         "Best weekend — sums unavailable across all days in each weekend (e.g. Fri+Sat+Sun)."
         " The weekend with the lowest total gets light green — most people can attend",
         fill_best, Font(name="Arial", bold=True, size=10, color="1B7A2B")),
        (LR+11, "Dark green",
         "Second best weekend — the next-best option if the best weekend doesn't work out",
         fill_second_best, Font(name="Arial", bold=True, size=10, color="FFFFFF")),
        (LR+12, "Red",
         "A critical person (\u2605) is unavailable during this weekend — not recommended",
         fill_wknd_bad, Font(name="Arial", bold=True, size=10, color="C00000")),
    ]
    for r, sym, desc, fill, fnt in legend:
        cs = ws.cell(row=r, column=1, value=sym)
        cs.font = fnt; cs.alignment = center; cs.border = thin
        if fill: cs.fill = fill
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        cd = ws.cell(row=r, column=2, value=desc)
        cd.font = normal_font; cd.alignment = left_a; cd.border = thin

    # ---------- CONFIG FOOTER ----------
    CFG = LR + 14
    ws.merge_cells(start_row=CFG, start_column=1, end_row=CFG, end_column=10)
    ws.cell(row=CFG, column=1,
            value=f"Config: Weekend = {weekend_label()} | "
                  f"Critical = {', '.join(CRITICAL)} | Year = {YEAR}").font = \
        Font(name="Arial", italic=True, size=9, color="888888")

import os as _os
OUT = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "event_planner.xlsx")
wb.save(OUT)
print(f"Saved → {OUT}")